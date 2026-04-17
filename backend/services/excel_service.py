"""Excel parsing and export helpers."""

from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from loguru import logger
from openpyxl import load_workbook

from utils.path_manager import get_output_dir


class ExcelService:
    """Handle Excel parsing, grouping, and export updates."""

    def __init__(self):
        self.base_required_columns = [
            '部門',
            '姓名',
            '計畫別',
            '出差日期時間（開始）',
            '出差日期時間（結束）',
        ]

    def _looks_like_address(self, text):
        text = str(text or '').strip()
        return any(token in text for token in ('市', '縣', '區', '鄉', '鎮', '路', '街', '大道', '號'))

    def _parse_destination_segment(self, segment):
        segment = str(segment or '').strip()
        if not segment:
            return None

        for separator in ('：', ':'):
            if separator in segment:
                name, address = segment.split(separator, 1)
                name = name.strip()
                address = address.strip()
                return {
                    '目的地名稱': name or address,
                    '終點地址': address or name,
                }

        dash_match = re.match(r'^(.{1,30}?)[\-－–—](.+)$', segment)
        if dash_match:
            name = dash_match.group(1).strip()
            address = dash_match.group(2).strip()
            if self._looks_like_address(address):
                return {
                    '目的地名稱': name or address,
                    '終點地址': address or name,
                }

        return {
            '目的地名稱': segment,
            '終點地址': segment,
        }

    def _split_destination_cell(self, value):
        text = str(value or '').strip()
        if not text:
            return []
        segments = [part.strip() for part in re.split(r'[;\uff1b\r\n]+', text) if part.strip()]
        if len(segments) == 1:
            city_prefix = (
                r'(?:\d{3,5})?'
                r'(?:\u53f0\u5317\u5e02|\u81fa\u5317\u5e02|\u65b0\u5317\u5e02|\u6843\u5712\u5e02|'
                r'\u53f0\u4e2d\u5e02|\u81fa\u4e2d\u5e02|\u53f0\u5357\u5e02|\u81fa\u5357\u5e02|'
                r'\u9ad8\u96c4\u5e02|\u57fa\u9686\u5e02|\u65b0\u7af9\u5e02|\u5609\u7fa9\u5e02|'
                r'\u65b0\u7af9\u7e23|\u82d7\u6817\u7e23|\u5f70\u5316\u7e23|\u5357\u6295\u7e23|'
                r'\u96f2\u6797\u7e23|\u5609\u7fa9\u7e23|\u5c4f\u6771\u7e23|\u5b9c\u862d\u7e23|'
                r'\u82b1\u84ee\u7e23|\u53f0\u6771\u7e23|\u81fa\u6771\u7e23|\u6f8e\u6e56\u7e23|'
                r'\u91d1\u9580\u7e23|\u9023\u6c5f\u7e23)'
            )
            implicit_segments = [part.strip() for part in re.split(r'\s+(?=' + city_prefix + r')', segments[0]) if part.strip()]
            if len(implicit_segments) > 1:
                segments = implicit_segments
        return [parsed for parsed in (self._parse_destination_segment(segment) for segment in segments) if parsed]

    def _expand_destination_rows(self, df):
        expanded_rows = []
        for source_index, (_, row) in enumerate(df.iterrows(), start=1):
            row_dict = row.to_dict()
            destination_text = row_dict.get('終點地址') or row_dict.get('目的地名稱')
            destinations = self._split_destination_cell(destination_text)
            if not destinations:
                row_dict['RouteChainGroup'] = f'row-{source_index}'
                row_dict['RouteChainIndex'] = 1
                row_dict['DisableChainedOrigin'] = 'Y'
                expanded_rows.append(row_dict)
                continue

            for index, destination in enumerate(destinations, start=1):
                new_row = row_dict.copy()
                new_row['目的地名稱'] = destination['目的地名稱']
                new_row['終點地址'] = destination['終點地址']
                new_row['行程序號'] = index
                new_row['RouteChainGroup'] = f'row-{source_index}'
                new_row['RouteChainIndex'] = index
                # Same-row multiple destinations are continuous; only the first leg starts from the fixed origin.
                new_row['DisableChainedOrigin'] = 'Y' if index == 1 else 'N'
                expanded_rows.append(new_row)

        return pd.DataFrame(expanded_rows)

    def _get_sort_key(self, date_value):
        if date_value is None:
            return datetime.min
        if isinstance(date_value, str):
            try:
                return pd.to_datetime(date_value)
            except Exception:
                return datetime.min
        if isinstance(date_value, pd.Timestamp):
            if pd.isna(date_value):
                return datetime.min
            return date_value.to_pydatetime()
        if isinstance(date_value, datetime):
            return date_value
        return datetime.min

    def _parse_datetime_value(self, value):
        if value is None or pd.isna(value):
            return pd.NaT
        if isinstance(value, (datetime, pd.Timestamp)):
            return pd.to_datetime(value, errors='coerce')

        text = str(value).strip()
        if not text:
            return pd.NaT

        period = None
        if '上午' in text:
            period = 'am'
            text = text.replace('上午', ' ')
        elif '下午' in text:
            period = 'pm'
            text = text.replace('下午', ' ')

        parsed = pd.to_datetime(text.strip(), errors='coerce')
        if pd.isna(parsed):
            return pd.NaT

        if period == 'pm' and parsed.hour < 12:
            parsed = parsed + pd.Timedelta(hours=12)
        elif period == 'am' and parsed.hour == 12:
            parsed = parsed - pd.Timedelta(hours=12)
        return parsed


    def parse_excel(self, file_path):
        try:
            logger.info(f"開始讀取 Excel 檔案: {file_path}")
            df = pd.read_excel(file_path, engine='openpyxl')
            logger.info(f"Excel 檔案讀取完成，共 {len(df)} 行, {len(df.columns)} 欄")

            column_mapping = {
                '部門': '部門',
                '姓名': '姓名',
                '員工姓名': '姓名',
                '計畫別': '計畫別',
                '專案': '計畫別',
                'ProjectName': '計畫別',
                '起點名稱': '起點名稱',
                '起點地址': '起點地址',
                'StartAddress': '起點地址',
                '出差日期時間（開始）': '出差日期時間（開始）',
                '出差起日': '出差日期時間（開始）',
                '出差日期時間（結束）': '出差日期時間（結束）',
                '出差迄日': '出差日期時間（結束）',
                '目的地名稱': '目的地名稱',
                '出差地點': '目的地名稱',
                '終點地址': '終點地址',
                'EndAddress': '終點地址',
                '連結': '連結',
                'Link': '連結',
            }
            rename_map = {src: dst for src, dst in column_mapping.items() if src in df.columns}
            df = df.rename(columns=rename_map)


            if '單位' in df.columns and '部門' not in df.columns:
                df['部門'] = df['單位']

            is_travel_table = '出差地點' in df.columns or '出差事由' in df.columns
            if is_travel_table:
                if '起點名稱' not in df.columns:
                    df['起點名稱'] = '安環高雄技術部'
                if '起點地址' not in df.columns:
                    df['起點地址'] = '813高雄市左營區博愛三路12號'
                df['DisableChainedOrigin'] = 'Y'
            elif '單位' in df.columns:
                if '起點名稱' not in df.columns:
                    df['起點名稱'] = df['單位']
                if '起點地址' not in df.columns:
                    df['起點地址'] = df['單位']

            missing_columns = [col for col in self.base_required_columns if col not in df.columns]
            if missing_columns:
                return {
                    'success': False,
                    'error': f"缺少必要欄位: {', '.join(missing_columns)}",
                    'data': None,
                }

            if '起點名稱' not in df.columns and '起點地址' not in df.columns:
                return {
                    'success': False,
                    'error': '缺少必要欄位: 起點名稱 或 起點地址',
                    'data': None,
                }

            if '目的地名稱' not in df.columns and '終點地址' not in df.columns:
                return {
                    'success': False,
                    'error': '缺少必要欄位: 目的地名稱 或 終點地址',
                    'data': None,
                }

            if '起點名稱' not in df.columns:
                df['起點名稱'] = df['起點地址']
            if '目的地名稱' not in df.columns:
                df['目的地名稱'] = df['終點地址']
            if '起點地址' not in df.columns:
                df['起點地址'] = df['起點名稱']
            if '終點地址' not in df.columns:
                df['終點地址'] = df['目的地名稱']


            df = self._expand_destination_rows(df)

            for date_col in ('出差日期時間（開始）', '出差日期時間（結束）'):
                if date_col in df.columns:
                    df[date_col] = df[date_col].apply(self._parse_datetime_value)

            for old_col in ('IsDriving', '是否自駕'):
                if old_col in df.columns:
                    df = df.drop(columns=[old_col])

            # 新上傳資料預設走開車模式。
            df['IsDriving'] = 'Y'
            df['OneWayKm'] = None
            df['RoundTripKm'] = None
            df['GoogleMapUrl'] = None
            df['StaticMapImage'] = None
            df['StepCount'] = None
            df['Polyline'] = None
            df['RouteSteps'] = None

            records = []
            for _, row in df.iterrows():
                record = {}
                for col in df.columns:
                    value = row[col]
                    if pd.isna(value):
                        record[col] = None
                    elif isinstance(value, pd.Timestamp):
                        record[col] = value.isoformat()
                    else:
                        record[col] = value
                records.append(record)

            logger.info(f"成功解析 Excel 檔案: {len(records)} 筆資料")
            return {
                'success': True,
                'data': records,
                'total_count': len(records),
            }
        except Exception as e:
            logger.error(f"解析 Excel 檔案錯誤: {str(e)}")
            return {
                'success': False,
                'error': f"解析 Excel 檔案失敗: {str(e)}",
                'data': None,
            }

    def group_by_project(self, records):
        try:
            grouped = {}
            for record in records:
                project_name = record.get('計畫別', '未分類')
                grouped.setdefault(project_name, []).append(record)

            for project_name in grouped:
                grouped[project_name].sort(
                    key=lambda item: self._get_sort_key(item.get('出差日期時間（開始）')),
                    reverse=False,
                )

            logger.info(f"成功分組: {len(grouped)} 個計畫別")
            return grouped
        except Exception as e:
            logger.error(f"分組錯誤: {str(e)}")
            return {}

    def add_calculation_results(self, file_path, records):
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            new_columns = [
                'OneWayKm',
                'RoundTripKm',
                'GoogleMapUrl',
                'StaticMapImage',
                'IsDriving',
                'StepCount',
                'Polyline',
                'RouteSteps',
            ]
            for col in new_columns:
                if col not in headers:
                    headers.append(col)
                    ws.cell(row=1, column=len(headers), value=col)

            col_index = {col: idx + 1 for idx, col in enumerate(headers)}

            for row_idx, record in enumerate(records, start=2):
                for key in new_columns:
                    value = record.get(key)
                    if value not in (None, ''):
                        ws.cell(row=row_idx, column=col_index[key], value=value)

            output_dir = get_output_dir()
            output_filename = f"updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = Path(output_dir) / output_filename
            wb.save(str(output_path))
            logger.info(f"成功產生更新 Excel 檔案: {output_path}")
            return str(output_path)
        except Exception as e:
            logger.error(f"更新 Excel 檔案錯誤: {str(e)}")
            raise
