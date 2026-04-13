"""Export routes for Excel and Word outputs."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import os
import zipfile

from flask import Blueprint, jsonify, request, send_file
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.excel_service import ExcelService
from services.word_service import WordService
from utils.log_sanitizer import sanitize_filename
from utils.path_manager import get_output_dir

bp = Blueprint('export', __name__)
excel_service = ExcelService()
word_service = WordService()


@bp.route('/excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json() or {}
        file_path = data.get('file_path')
        records = data.get('records', [])

        if not file_path or not os.path.exists(file_path):
            return jsonify({'status': 'error', 'message': '原始 Excel 檔案不存在'}), 400
        if not records:
            return jsonify({'status': 'error', 'message': '沒有可匯出的資料'}), 400

        output_path = excel_service.add_calculation_results(file_path, records)
        return send_file(
            output_path,
            as_attachment=True,
            download_name=os.path.basename(output_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        logger.error(f'匯出 Excel 失敗: {str(e)}')
        return jsonify({'status': 'error', 'message': f'匯出失敗: {str(e)}'}), 500


@bp.route('/word', methods=['POST'])
def export_word():
    try:
        data = request.get_json() or {}
        project_name = data.get('project_name', '未分類')
        records = data.get('records', [])
        fixed_origin = data.get('fixed_origin', '')

        if not records:
            return jsonify({'status': 'error', 'message': '沒有可匯出的資料'}), 400

        word_path = word_service.generate_report(project_name, records, fixed_origin)
        return send_file(
            word_path,
            as_attachment=True,
            download_name=os.path.basename(word_path),
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
    except Exception as e:
        logger.error(f'匯出 Word 失敗: {str(e)}')
        return jsonify({'status': 'error', 'message': f'匯出失敗: {str(e)}'}), 500


@bp.route('/word/batch', methods=['POST'])
def export_word_batch():
    try:
        data = request.get_json() or {}
        projects = data.get('projects') or {}
        fixed_origin = data.get('fixed_origin', '')

        grouped: dict[str, list[dict]] = {}
        if projects:
            for project_name, records in projects.items():
                if isinstance(records, list):
                    grouped[project_name or '未分類'] = [r for r in records if isinstance(r, dict)]
        else:
            records = data.get('records') or []
            for record in records:
                if not isinstance(record, dict):
                    continue
                project_name = record.get('計畫別') or record.get('project_name') or record.get('ProjectName') or '未分類'
                grouped.setdefault(project_name, []).append(record)

        if not grouped:
            return jsonify({'status': 'error', 'message': '沒有可匯出的資料'}), 400

        docx_paths: list[tuple[str, str]] = []
        for project_name, records in grouped.items():
            if not records:
                continue
            word_path = word_service.generate_report(project_name, records, fixed_origin)
            arc_name = sanitize_filename(f'{project_name}_里程報表.docx') or '未分類_里程報表.docx'
            docx_paths.append((word_path, arc_name))

        if not docx_paths:
            return jsonify({'status': 'error', 'message': '沒有成功產生任何 Word 檔'}), 500

        output_dir = get_output_dir()
        zip_filename = f"里程報表_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
        zip_path = Path(output_dir) / zip_filename

        used_names: set[str] = set()
        with zipfile.ZipFile(str(zip_path), 'w', zipfile.ZIP_DEFLATED) as zipf:
            for path, arc_name in docx_paths:
                name = arc_name
                if name in used_names:
                    stem, ext = os.path.splitext(name)
                    i = 2
                    while f'{stem}_{i}{ext}' in used_names:
                        i += 1
                    name = f'{stem}_{i}{ext}'
                used_names.add(name)
                zipf.write(path, name)

        logger.info(f'成功產生 ZIP 壓縮檔: {zip_path}, 包含 {len(docx_paths)} 個 Word 檔案')
        return send_file(zip_path, as_attachment=True, download_name=zip_filename, mimetype='application/zip')
    except Exception as e:
        logger.error(f'批次匯出 Word 失敗: {str(e)}')
        return jsonify({'status': 'error', 'message': f'匯出失敗: {str(e)}'}), 500


@bp.route('/template', methods=['GET'])
def export_template():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = '範本'

        headers = [
            '部門',
            '姓名',
            '計畫別',
            '起點名稱',
            '起點地址',
            '出差日期時間（開始）',
            '出差日期時間（結束）',
            '目的地名稱',
            '終點地址',
            '連結',
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 25

        example_data = [
            '安環處',
            '張三',
            'IDA智慧工安',
            '安環高雄處',
            '高雄市左營區博愛三路12號',
            '2024-10-22T09:00:00',
            '2024-10-22T17:00:00',
            '經濟部產業園區管理局',
            '高雄市楠梓區加昌路600號',
            '',
        ]
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for col, width in {
            'A': 15,
            'B': 12,
            'C': 20,
            'D': 20,
            'E': 28,
            'F': 22,
            'G': 22,
            'H': 24,
            'I': 28,
            'J': 72,
        }.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info('成功產生 Excel 範本檔案')

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='里程報表範本.xlsx',
        )
    except Exception as e:
        logger.error(f'產生 Excel 範本失敗: {str(e)}')
        return jsonify({'status': 'error', 'message': f'產生範本失敗: {str(e)}'}), 500
